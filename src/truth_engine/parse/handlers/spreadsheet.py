"""XLSX (openpyxl) and CSV (pandas) — the lossless content lives in
`StructuredTable`, never flattened into prose. `raw_text` is *also* set, but to
a deliberately lossy analysis-only projection (`tables_as_text`: filename +
sheet names + column headers + a bounded row sample), not the full table. Every
downstream analysis stage keys off `raw_text`, so a table with `raw_text=None`
is invisible to embeddings, NER, date resolution, phase assignment, and
direction/drift alike — for a spreadsheet-heavy corpus that means most of the
project never reaches the dashboard. The projection is what lets tables cluster,
get phased, and be dated by their filename/headers like any prose document.
"""

from __future__ import annotations

from pathlib import Path

from truth_engine.config import get_settings
from truth_engine.parse.handlers._common import header_from_row, json_safe, tables_as_text
from truth_engine.parse.registry import register
from truth_engine.parse.types import ParsedDocument, ParsedTable


def _read_xlsx(path: Path, *, data_only: bool) -> tuple[list[ParsedTable], list[str], bool]:
    """Read every sheet into `ParsedTable`s. Returns `(tables, sheet_names,
    had_header)` — `had_header` distinguishes a genuinely empty workbook from
    one whose *data* rows all came back empty (the formula-fallback trigger)."""
    import openpyxl  # lazy: pipeline extra

    workbook = openpyxl.load_workbook(path, data_only=data_only, read_only=True)
    try:
        tables: list[ParsedTable] = []
        had_header = False
        for sheet_name in workbook.sheetnames:
            rows_iter = workbook[sheet_name].iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row is None:
                continue
            had_header = True
            header = header_from_row(header_row)
            rows = [
                {h: json_safe(v) for h, v in zip(header, row, strict=False)}
                for row in rows_iter
                if row is not None and any(v is not None for v in row)
            ]
            tables.append(ParsedTable(source=sheet_name, table_schema=header, rows=rows))
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()
    return tables, sheet_names, had_header


@register("xlsx", "xlsm")
def parse_xlsx(path: Path) -> ParsedDocument:
    tables, sheet_names, had_header = _read_xlsx(path, data_only=True)
    # `data_only=True` returns each formula cell's *cached* value — but a
    # workbook generated programmatically and never opened in Excel has no
    # cached values, so every computed cell reads as None and the sheet looks
    # empty. When that happens (headers exist but zero data rows survived),
    # re-read with `data_only=False` to keep the formula text rather than
    # emitting a blank table.
    if had_header and not any(t.rows for t in tables):
        formula_tables, formula_sheets, _ = _read_xlsx(path, data_only=False)
        if any(t.rows for t in formula_tables):
            tables, sheet_names = formula_tables, formula_sheets

    settings = get_settings()
    return ParsedDocument(
        raw_text=tables_as_text(
            path.name,
            tables,
            max_rows_per_table=settings.parse_table_text_max_rows_per_table,
            max_chars=settings.parse_table_text_max_chars,
        ),
        structure={"sheet_names": sheet_names},
        embedded_metadata=None,
        tables=tables,
        parser_name="openpyxl",
        parser_version="2",
    )


@register("csv")
def parse_csv(path: Path) -> ParsedDocument:
    import pandas as pd  # lazy: pipeline extra

    # Real-world CSVs aren't all clean UTF-8: fall back to latin-1 (which
    # decodes any byte sequence) on a decode error rather than failing the
    # whole file, and skip ragged rows instead of raising on them.
    read_kwargs = dict(dtype=str, keep_default_na=False, on_bad_lines="warn")
    try:
        df = pd.read_csv(path, encoding="utf-8", **read_kwargs)
    except UnicodeDecodeError:
        df = pd.read_csv(path, encoding="latin-1", **read_kwargs)
    header = [str(c) for c in df.columns]
    table = ParsedTable(source=path.stem, table_schema=header, rows=df.to_dict(orient="records"))

    settings = get_settings()
    return ParsedDocument(
        raw_text=tables_as_text(
            path.name,
            [table],
            max_rows_per_table=settings.parse_table_text_max_rows_per_table,
            max_chars=settings.parse_table_text_max_chars,
        ),
        structure={"sheet_names": [path.stem]},
        embedded_metadata=None,
        tables=[table],
        parser_name="pandas",
        parser_version="2",
    )
